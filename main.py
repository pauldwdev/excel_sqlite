from sqlite3 import connect
from configparser import RawConfigParser
import classes
from glob import glob
import openpyxl
import os


#load configuration
config = RawConfigParser()
config.read('excel_sqlite.properties')
#print('loaded configuration file')

#generate database file name
databaseFile = config.get('database', 'database.path') + "/" + config.get('database', 'database.name')


#get excel files directory
docDir = config.get('documents', 'document.path')
processedDocumentPrefix = config.get('documents', 'document.read.prefix')
 



print('Options: 1. Load data into database OR 2. Reset db and files')
choice = input()
    
if choice == '1':

    #create connection to the database
    conn = connect(databaseFile)

    #open cursor
    cur = conn.cursor()
    excelDocList = glob(docDir + '/' + '[!_]*.xlsx')
    print('Processing xlsx files')
    for excelFile in excelDocList:
        wb = openpyxl.load_workbook(excelFile) 
        
        #loading population sheet
        populationSheet = wb['population']
        moreRows = True
        rowNumber = 2
        print('Processing file '+ excelFile)
        while moreRows:
            if populationSheet.cell(rowNumber,1).value != None:
                populationRecord = classes.populationRecord(populationSheet.cell(rowNumber, 1).value, populationSheet.cell(rowNumber, 2).value)
                cur.execute(populationRecord.createInsert())
            else:
                moreRows = False
            rowNumber = rowNumber + 1
            
        #loading food sheet
        foodSheet = wb['food']
        moreRows = True
        rowNumber = 2 
        while moreRows:
            if foodSheet.cell(rowNumber,1).value != None:
                foodRecord = classes.foodRecord(populationSheet.cell(rowNumber, 1).value, populationSheet.cell(rowNumber, 2).value)
                cur.execute(foodRecord.createInsert())
            else:
                moreRows = False
            rowNumber = rowNumber + 1
        print('Renaming files ')
        fileName = os.path.basename(excelFile)
        os.rename(excelFile, docDir + '\\' + processedDocumentPrefix + fileName)
    conn.commit()
    conn.close()
    print('Done loading data. Exiting.')
elif choice == '2':
    print('Renaming files so that they can be reprocessed...')
    excelDocList = glob(docDir + '/' + '_*.xlsx')
    
    for excelFile in excelDocList:
        fileName = os.path.basename(excelFile)
        print('file '+ excelFile + ' renaming to ' + docDir + '\\' + fileName[1:])
        os.rename(excelFile, docDir + '\\' + fileName[1:])
    #create connection to the database
    conn = connect(databaseFile)

    #open cursor
    cur = conn.cursor()
    print('Clearing pout previously loaded data from database...')
    cur.execute('delete from population')
    cur.execute('delete from food')
    conn.commit()
    conn.close()
    print('Done resetting files and data. Exiting')
else:
    print('wrong choice. Exiting')